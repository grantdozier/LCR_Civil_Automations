"""
Module A - Excel TOC Workbook Updater
Auto-populate Time of Concentration (TOC) calculation workbooks
"""
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from pathlib import Path
from typing import Dict, List, Optional
import logging

logger = logging.getLogger(__name__)


class TOCExcelUpdater:
    """
    Update Excel workbooks for Time of Concentration calculations.

    Handles:
    - Auto-population of drainage area data
    - Weighted C-value insertion
    - Area calculations (sqft and acres)
    - Formatting and formulas
    """

    def __init__(self, template_path: Optional[str] = None):
        """
        Initialize updater.

        Args:
            template_path: Path to Excel template (optional)
        """
        self.template_path = Path(template_path) if template_path else None
        self.workbook: Optional[Workbook] = None

    def load_template(self, template_path: str):
        """
        Load an existing Excel template.

        Args:
            template_path: Path to Excel file
        """
        self.template_path = Path(template_path)

        if not self.template_path.exists():
            raise FileNotFoundError(f"Template not found: {template_path}")

        try:
            self.workbook = load_workbook(self.template_path)
            logger.info(f"Loaded Excel template: {self.template_path.name}")
        except Exception as e:
            logger.error(f"Error loading Excel template: {e}")
            raise

    def create_new_workbook(self):
        """Create a new blank workbook"""
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)  # Remove default sheet

    def update_drainage_area_data(
        self,
        sheet_name: str,
        drainage_areas: List[Dict],
        start_row: int = 2
    ):
        """
        Update drainage area data in specified sheet.

        Args:
            sheet_name: Name of worksheet to update
            drainage_areas: List of drainage area dictionaries
                          Each dict should have: area_label, total_area_acres,
                          impervious_area_acres, pervious_area_acres, weighted_c_value
            start_row: Row number to start inserting data (default: 2, assuming row 1 is headers)
        """
        if not self.workbook:
            raise ValueError("No workbook loaded. Call load_template() or create_new_workbook() first")

        # Get or create sheet
        if sheet_name in self.workbook.sheetnames:
            ws = self.workbook[sheet_name]
        else:
            ws = self.workbook.create_sheet(sheet_name)
            self._create_headers(ws)

        current_row = start_row

        for area in drainage_areas:
            # Column A: Drainage Area Label
            ws.cell(row=current_row, column=1, value=area.get("area_label", ""))

            # Column B: Total Area (acres)
            ws.cell(row=current_row, column=2, value=area.get("total_area_acres", 0))

            # Column C: Impervious Area (acres)
            ws.cell(row=current_row, column=3, value=area.get("impervious_area_acres", 0))

            # Column D: Pervious Area (acres)
            ws.cell(row=current_row, column=4, value=area.get("pervious_area_acres", 0))

            # Column E: Weighted C-value
            ws.cell(row=current_row, column=5, value=area.get("weighted_c_value", 0))

            # Column F: Impervious %
            ws.cell(row=current_row, column=6, value=area.get("impervious_percentage", 0))

            # Apply formatting
            self._format_data_row(ws, current_row)

            current_row += 1

        logger.info(f"Updated {len(drainage_areas)} drainage areas in sheet '{sheet_name}'")

    def update_toc_calculations(
        self,
        sheet_name: str,
        toc_data: Dict[str, Dict],
        start_row: int = 2
    ):
        """
        Update Time of Concentration calculations.

        Args:
            sheet_name: Worksheet name
            toc_data: Dictionary mapping area_label to TOC data
                     Example: {
                         "E-DA1": {
                             "sheet_flow_length": 100,
                             "shallow_flow_length": 200,
                             "channel_flow_length": 500,
                             "tc_sheet": 5.2,
                             "tc_shallow": 3.8,
                             "tc_channel": 2.1,
                             "tc_total": 11.1
                         }
                     }
            start_row: Starting row for data
        """
        if not self.workbook:
            raise ValueError("No workbook loaded")

        if sheet_name not in self.workbook.sheetnames:
            ws = self.workbook.create_sheet(sheet_name)
            self._create_toc_headers(ws)
        else:
            ws = self.workbook[sheet_name]

        current_row = start_row

        for area_label, data in toc_data.items():
            ws.cell(row=current_row, column=1, value=area_label)
            ws.cell(row=current_row, column=2, value=data.get("sheet_flow_length", 0))
            ws.cell(row=current_row, column=3, value=data.get("shallow_flow_length", 0))
            ws.cell(row=current_row, column=4, value=data.get("channel_flow_length", 0))
            ws.cell(row=current_row, column=5, value=data.get("tc_sheet", 0))
            ws.cell(row=current_row, column=6, value=data.get("tc_shallow", 0))
            ws.cell(row=current_row, column=7, value=data.get("tc_channel", 0))

            # Total Tc formula (sum of components)
            ws.cell(row=current_row, column=8, value=f"=SUM(E{current_row}:G{current_row})")

            self._format_data_row(ws, current_row)
            current_row += 1

        logger.info(f"Updated TOC calculations for {len(toc_data)} areas")

    def save(self, output_path: str):
        """
        Save workbook to file.

        Args:
            output_path: Path for output Excel file
        """
        if not self.workbook:
            raise ValueError("No workbook to save")

        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        try:
            self.workbook.save(output_path)
            logger.info(f"Saved Excel workbook: {output_path}")
        except Exception as e:
            logger.error(f"Error saving workbook: {e}")
            raise

    def _create_headers(self, ws):
        """Create header row for drainage area sheet"""
        headers = [
            "Drainage Area",
            "Total Area (ac)",
            "Impervious (ac)",
            "Pervious (ac)",
            "Weighted C",
            "Impervious %"
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # Set column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 14
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12

    def _create_toc_headers(self, ws):
        """Create header row for TOC calculation sheet"""
        headers = [
            "Drainage Area",
            "Sheet Flow (ft)",
            "Shallow Flow (ft)",
            "Channel Flow (ft)",
            "Tc Sheet (min)",
            "Tc Shallow (min)",
            "Tc Channel (min)",
            "Tc Total (min)"
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # Set column widths
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws.column_dimensions[col].width = 15

    def _format_data_row(self, ws, row: int):
        """Apply formatting to a data row"""
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col in range(1, 9):  # Columns A-H
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

            # Format numbers
            if col > 1:  # Numeric columns
                cell.number_format = '0.00' if col <= 4 else '0.000'

    def get_sheet_data(self, sheet_name: str) -> List[Dict]:
        """
        Read data from a sheet.

        Args:
            sheet_name: Name of worksheet

        Returns:
            List of row dictionaries
        """
        if not self.workbook or sheet_name not in self.workbook.sheetnames:
            return []

        ws = self.workbook[sheet_name]
        data = []

        # Assuming row 1 is headers
        headers = [cell.value for cell in ws[1]]

        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(cell is None for cell in row):
                continue  # Skip empty rows

            row_dict = dict(zip(headers, row))
            data.append(row_dict)

        return data
